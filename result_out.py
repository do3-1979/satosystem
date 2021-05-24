#! python3

import csv
import os

import openpyxl
# 行列変換関数の取得
from openpyxl.utils import get_column_letter, column_index_from_string

while True:
    print('ファイルをドラッグ＆ドロップ >>>')
    input_file = input()
    # python内ではエスケープシーケンスでパスの修正が必要
    # 引数でもらう場合はいらない
    #input_file = ('C:\\python\\work\\bot_rt_07\\backtest\\result-partial-2021-01-24-00-40.csv')
    print('> ファイル名: ' + str(os.path.dirname(input_file)))
    
    # 入力ファイル
    input_obj = open(input_file,"r", newline = '', encoding="utf-8")
    input_reader = csv.reader(input_obj)
    
    input_data = list(input_reader) # csvはリストにして配列で参照が楽　ただしSEEKが進む
 
    # 出力ファイル
    tmp_path = str( os.path.splitext(input_file)[0] ) + '.xlsx'
    #print("tmp_path = {}".format(tmp_path))
    output_path = os.path.split(tmp_path)
    output_dir = output_path[0] #ディレクトリパス取り出し
    output_dir = os.path.join(output_dir + "\\result")
    #print("output_path = {}".format(output_path))

    # フォルダがなければ作る
    if os.path.exists(output_dir) != True:
        os.makedirs(output_dir)

    # 出力ファイル名
    output_file = os.path.join(output_dir + '\\' + output_path[1] ) # パスとファイル名結合

    tmp_rows = []

    # ファイルオブジェクトを先頭に戻す（忘れがち
    input_obj.seek(0)

    row_cnt = 0

    for row in input_reader:
        tmp_rows.append(row)
        print('\r' + '> 入力行数 : ' + str(row_cnt), end='') # 先頭上書き、改行なし
        row_cnt += 1
    print('')

    input_obj.close()

    # 元ファイルのタイトル行
    title = ['','開始', '終了', '時間軸', '買い期間', '売り期間', \
    'ボラティリティ期間', 'ストップ レンジ', 'トレードリスク', '分割回数', \
    '追加ポジション', 'ボラティリティ終値比', '注文lot数の下限',\
     'トレード回数', '勝率', '平均リターン', 'ドローダウン', 'PF', '最終損益']

    # 抜き出したいデータ
    pickup_title = ['買い期間', '売り期間','ボラティリティ期間', \
     'トレードリスク', '分割回数', 'ボラティリティ終値比', '注文lot数の下限',\
     'トレード回数', '勝率', '平均リターン', 'ドローダウン', 'PF', '最終損益']

    row_cnt = 0
    skip_cnt = 0

    # 出力ワークシートを作成
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    for row in tmp_rows:
        # 1行をリストで得る
        tmp = list(row)
        # 出力1行の初期化
        out_row = []

        #print("tmp len = {0}".format(len(tmp)))

        # 列を操作
        for i in range(len(pickup_title)):
            # 抜き出したいデータの列番号を得る
            #print("i = {0}".format(i))
            idx = title.index(pickup_title[i])
            #print("idx = {0}".format(idx))
            # 該当の列番号のデータを出力に追加
            out_row.append(tmp[idx])

        #print("out_row[] = {0}".format(out_row))

        # PFが0未満の行は不要
        # タイトル行は飛ばす
        if row_cnt > 0:
            idx = title.index('PF')
            if float(tmp[idx]) < 1.0:
                skip_cnt += 1
                continue

        # 出力データを書き出し
        # row, columun は1から始まる
        for i in range(len(out_row)):
            ws.cell(row=row_cnt+1, column=i+1).value = out_row[i]

        print('\r' + '> 出力行数 : ' + str(row_cnt), end='') # 先頭上書き、改行なし

        row_cnt += 1
    print('')

    print('> PF < 1.0 で削除した行数 : ' + str(skip_cnt))
    print('> 出力先 : ' + output_file)

    wb.save( output_file )

#    output_obj.close()

# EOF
