import os
import json
import zipfile
import pandas as pd
from datetime import datetime
import logging
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from bybit_exchange import BybitExchange
from config import Config
import pprint
import re

class Util:
    def extract_and_export_logs(self, log_directory, num_logs, output_excel_file, start_time=None, end_time=None):
        """
        指定された数の圧縮ログファイルを読み込み、エクセルファイルにデータとグラフを出力します。

        Args:
            log_directory (str): 圧縮ログファイルが格納されているディレクトリ
            num_logs (int): 読み込むログファイルの数
            output_excel_file (str): 出力するエクセルファイルの名前
            start_time (str): 開始時刻 (%Y/%m/%d %H:%M:%S 形式)
            end_time (str): 終了時刻 (%Y/%m/%d %H:%M:%S 形式)
        """
        if start_time and end_time:
            start_time = datetime.strptime(start_time, "%Y/%m/%d %H:%M")
            end_time = datetime.strptime(end_time, "%Y/%m/%d %H:%M")

        # 圧縮ログファイルのリストを作成
        log_files = []
        for root, _, files in os.walk(log_directory):
            for file in files:
                if file.endswith(".zip") or file.endswith(".json"):
                    log_files.append(os.path.join(root, file))

        # 新しいものから指定された数のログファイルを選択
        selected_log_files = sorted(log_files, reverse=True)[:num_logs]

        # データを格納するためのリスト
        data = []
        proccess_num_logs = min(num_logs, len(selected_log_files))

        # 選択されたログファイルからデータを抽出
        for i, log_file in enumerate(reversed(selected_log_files)):  # 逆順に処理
            print(f"Processing file {i + 1}/{proccess_num_logs}: {log_file}", end='\r')  # 処理中のファイルを表示
            if log_file.endswith(".zip"):
                with zipfile.ZipFile(log_file, "r") as log_zip:
                    with log_zip.open(log_zip.namelist()[0]) as log_json:
                        log_data = pd.read_json(log_json)
                        # データが期間内のものであれば追加
                        if start_time <= log_data['close_time'].max() and end_time >= log_data['close_time'].min():
                            data.append(log_data)
            elif log_file.endswith(".json"):
                with open(log_file, "r") as log_json:
                    log_data = pd.read_json(log_json)
                    # データが期間内のものであれば追加
                    if start_time <= log_data['close_time'].max() and end_time >= log_data['close_time'].min():
                        data.append(log_data)
        print(f"Processing file {i + 1}/{proccess_num_logs}: {log_file} completed")  # 処理中のファイルを表示

        # データを連結
        combined_data = pd.concat(data, ignore_index=True)

        # エクセルファイルに出力
        #output_excel_file_path = os.path.join(log_directory, output_excel_file)
        #workbook = Workbook()
        #writer = pd.ExcelWriter(output_excel_file_path, engine='openpyxl')
        #writer.book = workbook

        # カラムの選択
        selected_columns = [
            "real_time",
            "close_time_dt",
            "open_price",
            "high_price",
            "low_price",
            "close_price",
            "Volume",
            "stop_price",
            "position_price",
            "position_size",
            "position_quantity",
            "profit_and_loss",
            "total_profit_and_loss",
            "volatility",
            "stop_offset",
            "psar",
            "psarbull",
            "psarbear",
            "adx",
            "adx_bull",
            "adx_bear",
            "stop_psar_stop_offset",
            "stop_price_surge_stop_offset",
            "pvo_val",
            "decision",
            "side",
            "order_type",
            "dc_h",
            "dc_l",
            "donchian",
            "pvo",
            "positions"
        ]

        # 選択したカラムでデータをフィルタリング
        combined_data = combined_data[selected_columns]

        # 選択したカラムの順番に並べ替え
        combined_data = combined_data[[
            #"close_time_dt",
            "real_time",
            "high_price",
            "low_price",
            "close_price",
            "stop_price",
            "position_price",
            "dc_h",
            "dc_l",
            "Volume",
            "volatility",
            "decision",
            "side",
            "position_size",
            "position_quantity",
            "profit_and_loss",
            "total_profit_and_loss",
            "donchian",
            "pvo",
            "psar",
            "psarbull",
            "psarbear",
            "adx",
            "adx_bull",
            "adx_bear",
            "stop_offset",
            "stop_psar_stop_offset",
            "stop_price_surge_stop_offset",
        ]]

        # position_priceが0の場合、最小値に置き換え
        selected_columns = [
            "high_price",
            "low_price",
            "close_price",
            "dc_h",
            "dc_l",
        ]

        # min() を用いて最小値を求める
        min_position_price = combined_data[selected_columns].min().min()
        min_stop_price = min_position_price
        print(f"min_position_price = {min_position_price}")
        combined_data['position_price'] = combined_data['position_price'].replace(0, min_position_price)
        combined_data['stop_price'] = combined_data['stop_price'].replace(0, min_stop_price)

        config_dict = Config().to_dict()
        config_df = pd.DataFrame(list(config_dict.items()), columns=['Parameter', 'Value'])
        output_excel_file_path = os.path.join(log_directory, output_excel_file)
        with pd.ExcelWriter(output_excel_file_path, engine='openpyxl') as writer:
            # パラメータをエクセルに出力
            print("Generating Param sheet...", end='\r')
            config_df.to_excel(writer, index=False, sheet_name='Param', engine='openpyxl')
            print("Generating Param sheet...Done") 
            print("Generating Data sheet...", end='\r')
            # データをエクセルに出力
            combined_data.to_excel(writer, index=False, sheet_name='Data', engine='openpyxl')
            print("Generating Data sheet...Done") 

        # グラフをエクセルに出力
        print("Load file...", end='\r')
        workbook = load_workbook(output_excel_file_path)
        print("Load file...Done") 
        column_name = "value"
        print("Generating Chart sheet...", end='\r')
        chart_sheet = workbook.create_sheet(title="Chart")
        profit_and_loss_sheet = workbook.create_sheet(title="PandL")
        data_sheet = workbook['Data']
        self.generate_line_chart(combined_data, column_name, chart_sheet, data_sheet)
        print("Generating Chart sheet...Done")
        print("Generating Profit and Loss sheet...", end='\r')
        self.generate_line_profit_and_loss(combined_data, column_name, profit_and_loss_sheet, data_sheet)
        print("Generating Profit and Loss sheet...Done")
        # 既存の "Sheet" シートを削除
        #if "Sheet" in writer.book.sheetnames:
        #    std_sheet = writer.book["Sheet"]
        #    writer.book.remove(std_sheet)

        # エクセルファイルを保存
        print("File exporting...", end='\r')
        workbook.save(output_excel_file_path)
        print("File exporting...Completed!!")
        
    def generate_line_profit_and_loss(self, data, column_name, profit_and_loss_sheet, data_sheet):
        """
        データから指定されたカラムの折れ線グラフを生成し、指定されたシートに追加します。

        Args:
            data (DataFrame): グラフを生成するデータ
            column_name (str): グラフを生成するカラムの名前
            chart_sheet (Worksheet): グラフを追加するシート
        """
        chart = LineChart()
        chart.title = "損益履歴"
        # chart.style = 42
        chart.y_axis.title = column_name
        chart.x_axis.title = "real_time"  # x軸のタイトルを修正

        data_rows = list(dataframe_to_rows(data, index=False, header=True))
        real_time = Reference(data_sheet, min_col=1, min_row=2, max_row=len(data_rows)+1)
        profit_and_loss = Reference(data_sheet, min_col=15, min_row=1, max_row=len(data_rows)+1)
        total_profit_and_loss = Reference(data_sheet, min_col=16, min_row=1, max_row=len(data_rows)+1)

        chart.add_data(profit_and_loss, titles_from_data=True)
        chart.add_data(total_profit_and_loss, titles_from_data=True)

        chart.set_categories(real_time)

        # グラフの大きさを設定
        chart.width = 50  
        chart.height = 20  

        profit_and_loss_sheet.add_chart(chart, "A1")

    def generate_line_chart(self, data, column_name, chart_sheet, data_sheet):
        """
        データから指定されたカラムの折れ線グラフを生成し、指定されたシートに追加します。

        Args:
            data (DataFrame): グラフを生成するデータ
            column_name (str): グラフを生成するカラムの名前
            chart_sheet (Worksheet): グラフを追加するシート
        """
        chart = LineChart()
        chart.title = "取引履歴"
        #chart.style = 1
        chart.y_axis.title = column_name
        chart.x_axis.title = "real_time"  # x軸のタイトルを修正

        data_rows = list(dataframe_to_rows(data, index=False, header=True))
        
        real_time = Reference(data_sheet, min_col=1, min_row=2, max_row=len(data_rows)+1)
        high_price = Reference(data_sheet, min_col=2, min_row=1, max_row=len(data_rows)+1)
        low_price = Reference(data_sheet, min_col=3, min_row=1, max_row=len(data_rows)+1)
        close_price = Reference(data_sheet, min_col=4, min_row=1, max_row=len(data_rows)+1)
        stop_price = Reference(data_sheet, min_col=5, min_row=1, max_row=len(data_rows)+1)
        position_price = Reference(data_sheet, min_col=6, min_row=1, max_row=len(data_rows)+1)
        dc_h = Reference(data_sheet, min_col=7, min_row=1, max_row=len(data_rows)+1)
        dc_l = Reference(data_sheet, min_col=8, min_row=1, max_row=len(data_rows)+1)
        psar = Reference(data_sheet, min_col=19, min_row=1, max_row=len(data_rows)+1)

        #chart.add_data(real_time, titles_from_data=True)
        chart.add_data(high_price, titles_from_data=True)
        chart.add_data(low_price, titles_from_data=True)
        chart.add_data(close_price, titles_from_data=True)
        chart.add_data(stop_price, titles_from_data=True)
        chart.add_data(position_price, titles_from_data=True)
        chart.add_data(dc_h, titles_from_data=True)
        chart.add_data(dc_l, titles_from_data=True)
        chart.add_data(psar, titles_from_data=True)
        
        chart.set_categories(real_time)

        # TODO 生成したログファイルからデータ集約
        # TODO データのjsonファイル生成
        # TODO 各ログディレクトリからjsonファイルを集約 

        # グラフの大きさを設定
        chart.width = 50  # 幅を15に変更
        chart.height = 20  # 高さを10に変更

        # Y軸の最大値と最小値を指定
        selected_columns = [
            "high_price",
            "low_price",
            "close_price",
            "dc_h",
            "dc_l",
            "psar",
        ]

        # データ中の最小、最大を求める
        y_scale_max = data[selected_columns].max().max()
        y_scale_min = data[selected_columns].min().min()
        
        #chart.x_axis.scaling.min = x軸最小
        #chart.x_axis.scaling.max = x軸最大
        chart.y_axis.scaling.min = y_scale_min
        chart.y_axis.scaling.max = y_scale_max

        chart_sheet.add_chart(chart, "A1")

    def fetch_ohlcv_and_save_to_json(self, exchange, output_directory="ohlcv_json_data", output_filename="ohlcv_data.json"):
        start_epoch = Config.get_start_epoch()
        end_epoch = Config.get_end_epoch()
        #start_time = datetime.strptime(start_epoch, "%Y/%m/%d %H:%M")
        #end_time = datetime.strptime(end_epoch, "%Y/%m/%d %H:%M")
                
        print(f"start_epoch: {start_epoch} end_epoch: {end_epoch} ")
        ohlcv_data = exchange.fetch_ohlcv(start_epoch, end_epoch, 1)

        if not os.path.exists(output_directory):
            os.makedirs(output_directory)

        output_filepath = os.path.join(output_directory, output_filename)

        with open(output_filepath, 'w') as json_file:
            json.dump(ohlcv_data, json_file)
            

        print(f"OHLCVデータをJSONファイルに保存しました: {output_filepath}")

    def extract_parameters_and_results(self, log_directory, output_excel_file):
        # ログファイルの検索
        log_files = []
        for root, _, files in os.walk(log_directory):
            for file in files:
                if file.startswith("log_") and file.endswith(".txt"):
                    log_files.append(os.path.join(root, file))

        # ファイル名でソート
        log_files.sort()

        # 抽出するキーワード
        keywords = [
            "Entry Times:",
            "Entry Range:",
            "Stop Range:",
            "Stop AF:",
            "Stop AF Add:",
            "Stop AF Max:",
            "surge follow price ratio:",
            "Lot Limit Lower:",
            "Balance Tether Limit:",
            "Psar Time Frame:",
            "Volatility Term:",
            "Donchian Buy Term:",
            "Donchian Sell Term:",
            "PVO Short Term:",
            "PVO Long Term:",
            "PVO Threshold:",
            "最終損益:",
            "プロフィットファクター:",
            "最大ドローダウン:",
            "最大ドローダウン率:"
        ]

        """
            "Market:",
            "Market Unit:",
            "Start Time (epoch):",
            "End Time (epoch):",
            "Risk Percentage:",
            "Account Balance:",
            "Leverage:",
            "Server Retry Wait:",
            "Bot Operation Cycle:",
            "Back Test Mode:",
        """
        # エクセルファイルにデータを書き込む
        with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:
            data = {}
            for i, log_file in enumerate(log_files):
                # ファイル名からパス部分を取り除く
                file_name = os.path.basename(log_file)
                # ファイル名の列を追加
                keyword = "File Name"
                if keyword not in data:
                    data[keyword] = []
                data[keyword].append(file_name)

                print(f"Processing file {i+1}/{len(log_files)}: {file_name}", end='\r')
                with open(log_file, "r") as input_f:
                    for line in input_f:
                        for keyword in keywords:
                            if keyword in line:
                                # キーワード行だけを抽出
                                output_line = line.split(" - INFO - ", 1)[-1]
                                keyword, value = output_line.split(':', 1)
                                keyword = keyword.strip()  # 先頭と末尾の空白を削除
                                value = value.strip().split(' [')[0]      # 先頭と末尾の空白を削除し、[BTC/USD] や [%] を削除
                                if keyword not in data:
                                    data[keyword] = []
                                #print(f"file_name {file_name} keyword {keyword} value {value}")
                                data[keyword].append(value)

                        if "最終ポートフォリオ" in line:
                            # 正規表現を使用してquantity、side、position_priceの値を抽出
                            match = re.search(r"'quantity':\s*([0-9.]+),\s*'side':\s*'(\w+)',\s*'position_price':\s*([0-9.]+)", line)
                            if match:
                                quantity = match.group(1)
                                side = match.group(2)
                                position_price = match.group(3)
                                if "quantity" not in data:
                                    data["quantity"] = []
                                if "side" not in data:
                                    data["side"] = []
                                if "position_price" not in data:
                                    data["position_price"] = []
                                data["quantity"].append(quantity)
                                data["side"].append(side)
                                data["position_price"].append(position_price)

            # 最長のリストの長さに合わせて、他のリストに空の文字列を追加する
            max_length = max(len(lst) for lst in data.values())
            for key in data:
                data[key] += [''] * (max_length - len(data[key]))

            # データをエクセルに書き込む
            df = pd.DataFrame(data)
            df.to_excel(writer, index=False, sheet_name="Log Data")

            print("\nExporting to Excel completed!")

if __name__ == "__main__":
    # Loggerクラスの初期化
    util = Util()

    log_directory = "logs"  # ログファイルのディレクトリ
    #log_directory = "logs_0000020"  # ログファイルのディレクトリ
    #log_directory = "../test/test_data"  # ログファイルのディレクトリ

    num_logs_to_read = 400  # 読み込むログファイルの数
    output_excel_file = "combined_logs.xlsx"  # 出力エクセルファイルの名前
    
    #----------------
    # グラフ化機能使用
    
    start_time = "2025/5/1 1:00"  # 開始時刻 (例: "2023/01/01 00:00")
    end_time = "2025/5/24 18:00"    # 終了時刻 (例: "2023/01/02 00:00")
    #start_time = Config.get_start_time()
    #end_time = Config.get_end_time()

    util.extract_and_export_logs(log_directory, num_logs_to_read, output_excel_file, start_time, end_time)

    #---------------------
    # バックテスト集約使用例
    #log_directory = "."  # ログファイルのディレクトリ
    #output_bt_excel_file = "backtest_logs.xlsx"  # 出力エクセルファイルの名前

    # パラメータと計算結果を抽出してエクセルファイルに出力
    #util.extract_parameters_and_results(log_directory, output_bt_excel_file)

    # 1分足の指定ログ取得
    # exchange = BybitExchange(Config.get_api_key(), Config.get_api_secret())
    # util.fetch_ohlcv_and_save_to_json(exchange)